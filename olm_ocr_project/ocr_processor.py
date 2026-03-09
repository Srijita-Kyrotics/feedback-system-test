import os
import argparse
import json
import pandas as pd
from PIL import Image 
from PIL import Image
import torch
#import statements
# Heavy imports moved inside functions or made optional
# from transformers import Qwen2VLForConditionalGeneration, AutoTokenizer, AutoProcessor
# from qwen_vl_utils import process_vision_info

def load_model(model_name="Qwen/Qwen2-VL-7B-Instruct"):
    """
    Load the Qwen2-VL model and processor.
    """
    import torch
    from transformers import Qwen2VLForConditionalGeneration, AutoProcessor, BitsAndBytesConfig
    print(f"Loading model: {model_name} (with 4-bit quantization)...")
    try:
        quantization_config = BitsAndBytesConfig(
            load_in_4bit=True,
            bnb_4bit_compute_dtype=torch.float16,
            bnb_4bit_quant_type="nf4",
            bnb_4bit_use_double_quant=True,
        )
        model = Qwen2VLForConditionalGeneration.from_pretrained(
            model_name,
            torch_dtype=torch.float16,
            device_map="auto",
            quantization_config=quantization_config,
        )
        processor = AutoProcessor.from_pretrained(model_name)
        return model, processor
    except Exception as e:
        print(f"Error loading model: {e}")
        print("Ensure you have a GPU available or sufficient CPU RAM. You might need to install 'accelerate'.")
        raise

# ... (imports)

def update_excel_template(template_path, output_path, data):
    """
    Load template, fill data, and save to output_path.
    """
    import openpyxl
    import re
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # 1. Update Header (Cell A1)
        # Construct the header string. 
        # We try to preserve original format or reconstruct it.
        # "Department: {Department}\nSemester: {Semester}..."
        
        header_text = ws['A1'].value or ""
        
        header_fields_map = {
            "Department": data.get("Department", ""),
            "Semester": data.get("Semester", ""),
            "Year": data.get("Year", ""),
            "Course Code": data.get("CourseCode", ""),
            "Course Name": data.get("CourseName", ""),
            "Course Taught By": data.get("TaughtBy", ""),
            r"\(Full/Part\)": data.get("FullPart", "") 
        }

        # Create a combined pattern of all keys to look ahead
        # Sort keys by length descending to match longest first if there's overlap, 
        # though here they are distinct.
        all_keys = list(header_fields_map.keys())
        # Escape keys for regex (except the one already escaped)
        escaped_keys = [k if k.startswith(r"\(") else re.escape(k) for k in all_keys] 
        lookahead_pattern = "|".join(escaped_keys) + r"|$|\n"

        new_header = header_text
        for key, val in header_fields_map.items():
            # Current key pattern: (Key\s*:\s*)
            # Value pattern: (.*?)(?=(NextKey:|OtherKey:|...|$|\n))
            # We construct the regex dynamically for each key
            
            # Key to match current field
            key_regex = key if key.startswith(r"\(") else re.escape(key)
            
            # The value part matches anything non-greedy until it sees another key or end of line
            # logic: find (Key:\s*)(value)(?=(any_other_key_start))
            
            # We need to look ahead for "Key:\s*" or "Key:"?
            # actually just the key name is enough as lookahead
            
            pattern = rf"({key_regex}\s*:\s*)(.*?)(?={lookahead_pattern})"
            
            # But wait, lookahead_pattern includes ALL keys, including the current one.
            # We don't want to stop at the current key (start), but that's fine because we consumed it in group 1.
            # We want to stop at the START of the NEXT key.
            # Since we iterate, we might replace part of the string.
            # It's safer to perform replacements?
            # If we replace "Year:" with "Year: 2026", does that break "Semester: ...Year:"?
            # No, we assume keys are preserved.
            
            # Refined Approach:
            # use a simpler robust regex: stop at specific known boundary chars if feasible, 
            # or use the specific known next keys for each field if order is fixed.
            # Given the text: "Semester: 7thYear: 2025..."
            # Value for Semester is "7th". Next key is "Year".
            # Value for "Course Taught By" is "Dr.Rumi Sen". Next key is "(Full/Part)".
            
            # Let's use the all_keys lookahead.
            # Exclude current key from lookahead to avoid 0-length match if that was possible? 
            # No.
            
            # One issue: "Course Name" contains "Course". "Course Code" contains "Course".
            # "Course Taught By" contains "Course".
            # If we look ahead for "Course", we might stop early?
            # "Course Code" vs "Course Name".
            # Lookahead should match full keys.
            
            # Let's try to match: ({key}\s*:\s*)(.*?)(?=(?:{lookahead_pattern}))
            # re.sub might get complicated if we do it iteratively and text changes?
            # Text grows. "Value" -> "NewValue".
            # Keys remain.
            
            def replacer(match):
                return f"{match.group(1)}{val}"
            
            new_header = re.sub(pattern, replacer, new_header, count=1, flags=re.IGNORECASE)

        ws['A1'] = new_header
        
        # 2. Update Answers (Rows 4 to 18 -> Indices 4 to 18 in Excel is Row 4..18)
        # In openpyxl, rows are 1-indexed.
        # Check previous analysis: 
        # Row 0 (pandas) = Row 1 (Excel) -> Header
        # Row 3 (pandas) = Row 4 (Excel) -> Q1
        
        # Map Q1..Q14 to cells C4..C18 (Column 3)
        for i in range(1, 15):
            q_key = f"Q{i}"
            val = data.get(q_key, "")
            # Row index matches Q number + 3? 
            # Q1 -> Row 4. Q14 -> Row 17?
            # Let's check: 3(pandas)=4(excel). 17(pandas)=18(excel).
            # Wait, 17-3 = 14 items.
            # Q1 at row 4. Q14 at row 17.
            # cell = ws.cell(row=i+3, column=3)
            # wait, if Q1 is row 4. Q14 is row 17.
            # (17-4)+1 = 14. Correct.
            ws.cell(row=i+3, column=3).value = val
            
        # 3. Update Q15
        # Assuming Q15 answer goes to a specific cell.
        # Let's append to cell A21 (Row 21) or just below the Q15 text.
        # Using Row 21, Col 1 just to be safe.
        ws.cell(row=21, column=1).value = data.get("Q15", "")
        
        wb.save(output_path)
        print(f"Saved populated Excel to {output_path}")
        return True
        
    except Exception as e:
        print(f"Error updating Excel template: {e}")
        return False

def process_folder(folder_path, model, processor, template_path):
    """
    Process a single folder containing images.
    """
    images = []
    # ... (image finding logic)
    valid_extensions = ('.jpg', '.jpeg', '.png', '.bmp', '.tiff')
    for f in os.listdir(folder_path):
        if f.lower().endswith(valid_extensions):
            full_path = os.path.join(folder_path, f)
            images.append(full_path)

    if not images:
        print(f"No images found in {folder_path}")
        return None

    # Custom sort key
    def sort_key(path):
        filename = os.path.basename(path)
        base, _ = os.path.splitext(filename)
        try:
            return int(base)
        except ValueError:
            return filename  
    images.sort(key=sort_key)
    
    print(f"Processing {len(images)} images in {folder_path}...")
    
    from qwen_vl_utils import process_vision_info
    
    # Process images one by one to save memory, then combine (or just take the first if it's a single form)
    # The current pipeline seems to expect one JSON for the whole folder.
    # Usually these are multi-page forms of the SAME survey.
    
    # We will try to combine them in the prompt or process them and merge JSONs.
    # For now, let's just try processing them one by one and merging the data.
    aggregated_data = {}
    
    # Prompt
    prompt_text = """Analyze this handwritten Aliah University Survey form carefully:
    1. HEADER: Identify the handwritten text. Specifically look for:
       - Department (e.g., "Nursing")
       - Semester (e.g., "7th")
       - Year (e.g., "2025")
       - Taught By (e.g., "Dr. Rumi Sen")
       - Course Name (e.g., "B.Sc Nursing")
    2. SCORES (Q1-Q14): The form has a table with columns 1 to 5. Identify the column with the handwritten 'tick' (check mark) for each question.
       - Output the numeric score (1-5) for each question.
    3. JSON OUTPUT: Return only a flat JSON object with keys: "Department", "Semester", "Year", "CourseCode", "CourseName", "TaughtBy", "FullPart", "Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7", "Q8", "Q9", "Q10", "Q11", "Q12", "Q13", "Q14", "Q15".
    
    If text is cursive, read it carefully (e.g., "Nursing", "Rumi Sen"). If a field is blank on one page, it may be on the other image.
    """

    for img_path in images:
        print(f"  Processing image: {os.path.basename(img_path)}")
        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "image", "image": img_path},
                    {"type": "text", "text": prompt_text}
                ],
            }
        ]

        text = processor.apply_chat_template(
            messages, tokenize=False, add_generation_prompt=True
        )
        
        image_inputs, video_inputs = process_vision_info(messages)
        inputs = processor(
            text=[text],
            images=image_inputs,
            videos=video_inputs,
            padding=True,
            return_tensors="pt",
        )
        inputs = inputs.to(model.device)

        with torch.no_grad():
            generated_ids = model.generate(**inputs, max_new_tokens=1024)
        
        generated_ids_trimmed = [
            out_ids[len(in_ids) :] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
        ]
        output_text = processor.batch_decode(
            generated_ids_trimmed, skip_special_tokens=True, clean_up_tokenization_spaces=False
        )[0]
        
        # Robust JSON extraction
        try:
            # Find the first '{' and last '}' to extract the JSON block
            start_idx = output_text.find('{')
            end_idx = output_text.rfind('}')
            
            if start_idx != -1 and end_idx != -1:
                json_str = output_text[start_idx:end_idx+1]
                data = json.loads(json_str)
                
                # Exclusion list for non-data strings
                null_values = ["", "0", "None", "not provided", "none", "unknown", "n/a", "null", "not specified"]
                
                print(f"      Extracted {len(data)} fields from image.")
                
                for k, v in data.items():
                    val_str = str(v).strip().lower()
                    if val_str not in null_values:
                        aggregated_data[k] = v
            else:
                print(f"      Warning: No JSON block found in AI response for {img_path}")
                    
        except Exception as e:
            print(f"    Warning: Failed to parse or merge result for {img_path}: {e}")
            print(f"    Raw preview: {output_text[:200]}...")

    print(f"  Final Aggregated Teacher Data for {folder_path}:")
    print(f"    Teacher: {aggregated_data.get('TaughtBy', 'Unknown')}")
    print(f"    Dept: {aggregated_data.get('Department', 'Unknown')}")
    return json.dumps(aggregated_data)

def main():
    parser = argparse.ArgumentParser(description="Extract text from form images to Excel using OLM OCR (Qwen2-VL).")
    parser.add_argument("--input_dir", type=str, required=True, help="Root directory containing subfolders with images.")
    parser.add_argument("--model", type=str, default="Qwen/Qwen2-VL-7B-Instruct", help="HuggingFace model ID.")
    parser.add_argument("--template", type=str, default="template.xlsx", help="Path to Excel template.")
    parser.add_argument("--mock", action="store_true", help="Run in mock mode.")
    args = parser.parse_args()

    # Load model once
    model = None
    processor = None
    if not args.mock:
        try:
            model, processor = load_model(args.model)
        except Exception as e:
            print(f"Failed to load model: {e}")
            return
    else:
        print("Running in MOCK mode.")

    # Iterate over subdirectories
    for root, dirs, files in os.walk(args.input_dir):
        base_name = os.path.basename(root)
        # Skip hidden folders or non-content folders if needed
        
        has_images = any(f.lower().endswith(('.jpg', '.jpeg', '.png')) for f in files)
        
        if has_images:
            print(f"Processing folder: {root}")
            try:
                if args.mock:
                    # Mock output with all fields
                    extracted_text = json.dumps({
                        "Department": "Nursing",
                        "Semester": "7th",
                        "Year": "2025",
                        "CourseCode": "NURS101",
                        "CourseName": "B.Sc Nursing",
                        "TaughtBy": "Dr. Mock",
                        "FullPart": "Full",
                        "Q1": "5", "Q2": "4", "Q3": "5", "Q4": "5", "Q5": "5",
                        "Q6": "4", "Q7": "5", "Q8": "5", "Q9": "5", "Q10": "4",
                        "Q11": "5", "Q12": "5", "Q13": "5", "Q14": "4",
                        "Q15": "More practical sessions."
                    })
                else:
                    extracted_text = process_folder(root, model, processor, args.template)
                
                if extracted_text:
                    # Parse JSON
                    try:
                        cleaned_text = extracted_text.strip()
                        if cleaned_text.startswith("```json"):
                            cleaned_text = cleaned_text[7:]
                        if cleaned_text.endswith("```"):
                            cleaned_text = cleaned_text[:-3]
                        
                        data = json.loads(cleaned_text.strip())
                        
                        # Generate Excel using Template
                        output_filename = f"{base_name}.xlsx" if base_name else "extracted_data.xlsx"
                        output_path = os.path.join(root, output_filename)
                        
                        update_excel_template(args.template, output_path, data)
                        
                    except json.JSONDecodeError as e:
                        print(f"Failed to parse JSON: {e}")
                        with open(os.path.join(root, "raw_output.txt"), "w") as f:
                            f.write(extracted_text)
            except Exception as e:
                print(f"Error processing {root}: {e}")

if __name__ == "__main__":
    main()
